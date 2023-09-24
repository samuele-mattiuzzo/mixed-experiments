from datetime import datetime
from openpyxl import load_workbook

from models_openpyxl import MAPPING as M
from models_openpyxl import Product, Review

MIN_ROW = 2
MAX_ROW = 1000
READ_ONLY = True
VALUES_ONLY = True

existing_products_id = []
products = {}
reviews = []

def parse_review(row, product):
    parsed_date = datetime.strptime(
        row[M["RE_DATE"]], "%Y-%m-%d")

    return Review(id=row[M["RE_ID"]],
                    customer_id=row[M["RE_CUSTOMER"]],
                    stars=row[M["RE_STARS"]],
                    headline=row[M["RE_HEADLINE"]],
                    body=row[M["RE_BODY"]],date=parsed_date,
                    product=product)

def parse_product(row):
    return Product(id=row[M["PR_ID"]],
                    parent=row[M["PR_PARENT"]],
                    title=row[M["PR_TITLE"]],
                    category=row[M["PR_CATEGORY"]])

if __name__ == "__main__":
    workbook = load_workbook(filename="reviews-sample.xlsx", read_only=READ_ONLY)
    sheet = workbook.active

    print(sheet.title)
    print("")

    for row in sheet.iter_rows(min_row=MIN_ROW, 
                               max_row=MAX_ROW, 
                               values_only=VALUES_ONLY):

        pr_id = row[M["PR_ID"]]
        if pr_id in existing_products_id:
            product = products[pr_id]
        else:
            product = parse_product(row)
            products[product.id] = product
        
        review = parse_review(row, product)                
        reviews.append(review)


    for rev in sorted(reviews, key=lambda x: x.product.id):
        print(rev.stars, rev.product.id)
